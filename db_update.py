import openpyxl
#import pymysql
import sqlite3

print('Starting...')

conn = sqlite3.connect('star_interactive.db')
cur = conn.cursor()

inventory_file = "Dec 12 Inventory.xlsx" 
wb = openpyxl.load_workbook(inventory_file)
sheet = wb.active

total_items = sheet.max_row - 1
added_items = 0

added_skus = set()
failed_skus = list()
no_upc = set()

for row in range(2, sheet.max_row + 1):
	try:
		item_num = sheet['A' + str(row)].value
		item_sku = sheet['B' + str(row)].value
		item_desc = sheet['C' + str(row)].value
		item_inv = sheet['E' + str(row)].value
		item_upc = sheet['I' + str(row)].value
		item_info = (item_num, item_sku, item_desc, item_inv, item_upc)
		if item_upc == '':
			item_upc = str(0)
			no_upc.add(item_sku)
		if item_sku not in added_skus:
			cur.execute("INSERT INTO inventory VALUES (?, ?, ?, ?, ?)", item_info)
			conn.commit()
			print('Added', item_sku)
			added_skus.add(item_sku)
			added_items += 1
	except Exception as e:
		print(e)
		failed = (item_sku, e)
		failed_skus.append(failed)


# Close MySQL Database
cur.close()
conn.close()
print('Done, {} out of {} total items successfully added to database.'.format(added_items, total_items))
print('Number of SKUs with no associated UPC: ', str(len(no_upc)))
print('Failed SKUs:')
if len(failed_skus) > 0:
	for sku in failed_skus:
		print(sku[0], sku[1])