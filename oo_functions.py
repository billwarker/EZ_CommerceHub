import os
import datetime
import openpyxl
#import pymysql
from oo_dicts import *
import csv

def _grab_skus_upc(row, output_sheet):
	product_desc = output_sheet['AM'+ str(row)].value
	#print(product_desc)
	if product_desc.endswith('(14 Pack)'):
		# UPC
		output_sheet['CQ'+ str(row)] = '743724486834'
		# SKU
		output_sheet['CR'+ str(row)] = 'AO-14'

	elif product_desc.endswith('(6 Pack)'):
		# UPC
		output_sheet['CQ'+ str(row)] = '743724486827'
		# SKU
		output_sheet['CR'+ str(row)] = 'AO-6'

	elif product_desc.endswith('Set'):
		# UPC
		output_sheet['CQ'+ str(row)] = '743724487237'
		# SKU
		output_sheet['CR'+ str(row)] = 'AO-8'

	elif product_desc.startswith('AD-200'):
		# UPC
		output_sheet['CQ'+ str(row)] = 'IGNORE ME'
		# SKU
		output_sheet['CR'+ str(row)] = 'AD-200'

	elif product_desc.startswith('S330'):
		# UPC
		output_sheet['CQ'+ str(row)] = 'IGNORE ME'
		# SKU
		output_sheet['CR'+ str(row)] = 'S330'

	elif product_desc.startswith('S625'):
		# UPC
		output_sheet['CQ'+ str(row)] = 'IGNORE ME'
		# SKU
		output_sheet['CR'+ str(row)] = 'S625'

	elif product_desc.startswith('OI100R'):
		# UPC
		output_sheet['CQ'+ str(row)] = 'IGNORE ME'
		# SKU
		output_sheet['CR'+ str(row)] = 'OI-100R'

	elif product_desc.startswith('OI200'):
		# UPC
		output_sheet['CQ'+ str(row)] = 'IGNORE ME'
		# SKU
		output_sheet['CR'+ str(row)] = 'OI-200'

def _order_dates(row, output_sheet):
	today = datetime.date.today()
	tomorrow = datetime.date.today() + datetime.timedelta(days=1)
	output_sheet['BH'+ str(row)] = today.strftime("%m-%d-%Y")
	output_sheet['BN'+ str(row)] = tomorrow.strftime("%m-%d-%Y")

def _check_errors(row, final_col, output_sheet, error_rows):
	for col in range(1, final_col):
		col_letter = openpyxl.cell.cell.get_column_letter(col)
		if (output_sheet[col_letter + str(row)].value == 'IGNORE ME' or output_sheet[col_letter + str(row)].value == 'N/A'
			or output_sheet[col_letter + str(row)].value == '0'):
			error_rows.add(row)

def _mysql_lookup(row, output_sheet, cur):
	row_sku = output_sheet["CR" + str(row)].value
	try:
		cur.execute("SELECT item_upc FROM inventory WHERE item_sku = ?;",
			(row_sku,))
		row_upc = cur.fetchone()[0]
		#print(query.format(row_sku))
		print('Grabbed UPC from database for', row_sku)
		output_sheet["CQ" + str(row)] = str(row_upc)
	except Exception as e:
		print(e)

def process_sheet(wb_file, final_col, output_sheet, vendor_dict,
	offset, cur, error_rows, groupon_true=False, commerce_true=False):
	input_sheet = _csv_check(wb_file)
	
	if commerce_true:
		input_sheet = _commerce_filter(input_sheet)

	last_row = input_sheet.max_row
	print('SKUs:', last_row - 1)
	for row in range(2, last_row + 1):
		for col in range(1, final_col):
			try:
				col_letter = openpyxl.cell.cell.get_column_letter(col)
				if (vendor_dict[col_letter] == None
				 or vendor_dict[col_letter] == NA or vendor_dict[col_letter] == 'US'
				 or vendor_dict[col_letter] == 'NEED DATE'
				 or vendor_dict[col_letter] == 'Canada Post - Expedited Parcel'
				 or vendor_dict[col_letter] == 'ATS/TForce Integrated Solutions'
				 or vendor_dict[col_letter] == 'IGNORE ME'
				 or vendor_dict[col_letter] == 'Groupon' or vendor_dict[col_letter] == 'Staples'):
					output_sheet[col_letter + str(row + offset)] = vendor_dict[col_letter]
				else:
					output_sheet[col_letter + str(row + offset)] = str(input_sheet[vendor_dict[col_letter] + str(row)].value)
			except Exception as e:
				print(e)
		_order_dates((row + offset), output_sheet)
		
		if groupon_true:
			_grab_skus_upc(row + offset, output_sheet)

		_mysql_lookup((row + offset), output_sheet, cur)
		_check_errors((row + offset), final_col, output_sheet, error_rows)

	offset += last_row - 1

	return output_sheet, offset, error_rows

def _csv_check(file):
	if file.endswith('.csv'):
		print('Converting {} to .xlsx format...'.format(file))
		file_name = file.split('.csv')[0]
		wb = openpyxl.Workbook()
		sheet = wb.active

		with open(file, 'r') as f:
			reader = csv.reader(f)
			for r, row in enumerate(reader):
				for c, col in enumerate(row):
						cell = sheet.cell(row=r+1, column=c+1)
						cell.value = col
		return sheet
	else:
		wb = openpyxl.load_workbook(file)
		sheet = wb.active

		return sheet

### CommerceHub Processing

def _commerce_filter(input_sheet):
	print('Filtering out acknowledged, invoiced, and closed orders...')
	wb = openpyxl.Workbook()
	filtered_sheet = wb.active
	filtered_row = 2
	count = 0

	for row in range(input_sheet.min_row, input_sheet.max_row + 1):
		line_status = input_sheet['AY' + str(row)].value
		status = input_sheet['CI' + str(row)].value
		sub_status = input_sheet['CJ' + str(row)].value
		
		if line_status == 'new':
			if status == 'undelivered' and sub_status == 'undelivered':
				for col in range(1, input_sheet.max_column + 1):
					col_letter = openpyxl.cell.cell.get_column_letter(col)
					filtered_sheet[col_letter + str(filtered_row)] = input_sheet[col_letter + str(row)].value
				filtered_row += 1
				count += 1
	print('Found {} open orders.'.format(count))
	return filtered_sheet