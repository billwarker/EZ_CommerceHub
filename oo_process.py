import openpyxl
import csv
import time
from oo_dicts import *
from oo_functions import *
import datetime
import sys
import sqlite3
import os
from oo_settings import *

def process_output(groupon_file, commerce_file, staples_file, commerce2_file):
	# connect to db for SKUs and UPCs
	conn = sqlite3.connect(DATABASE)
	cur = conn.cursor()

	# Main Output Sheet
	output_wb = openpyxl.Workbook()
	output_sheet = output_wb.active

	# Star Interactive Sheet
	star_wb = openpyxl.Workbook()
	star_sheet = star_wb.active
	star_row = 2	# First row after header columns

	# SBW Sheet
	sbw_wb = openpyxl.Workbook()
	sbw_sheet = sbw_wb.active
	sbw_row = 2		# First row after header columns

	# write columns for new output wb, set col. width
	col_width = 20
	format_file = "oo_formatSheet.xlsx" # Always use this spreadsheet as the template for the column names/formatting
	format_wb = openpyxl.load_workbook(format_file)
	format_sheet = format_wb.active
	for col in range(1, format_sheet.max_column + 1):
		col_letter = openpyxl.cell.cell.get_column_letter(col)
		
		output_sheet[col_letter + str(1)] = format_sheet[col_letter + str(1)].value
		output_sheet.column_dimensions[col_letter].width = col_width

		star_sheet[col_letter + str(1)] = format_sheet[col_letter + str(1)].value
		star_sheet.column_dimensions[col_letter].width = col_width

		sbw_sheet[col_letter + str(1)] = format_sheet[col_letter + str(1)].value
		sbw_sheet.column_dimensions[col_letter].width = col_width

	final_col = format_sheet.max_column + 1

	# row offset for multiple input sheets
	offset = 0
	# error_rows for final checking
	error_rows = set()

	if commerce_file:
		print('Adding CommerceHub')
		output_sheet, offset, error_rows = process_sheet(commerce_file, final_col, output_sheet, commercehub_dict,
			offset, cur, error_rows, commerce_true=True)

	if commerce2_file:
		print('Adding CommerceHub (CA)')
		output_sheet, offset, error_rows = process_sheet(commerce2_file, final_col, output_sheet, commercehub2_dict,
			offset, cur, error_rows, commerce_true=True)


	if groupon_file:
		print('Adding Groupon')
		output_sheet, offset, error_rows = process_sheet(groupon_file, final_col, output_sheet, groupon_dict,
			offset, cur, error_rows, groupon_true=True)

	if staples_file:
		print('Adding Staples')
		output_sheet, offset, error_rows = process_sheet(staples_file, final_col, output_sheet, staples_dict,
			offset, cur, error_rows)

	# Writing output XLSX files
	today = datetime.date.today().strftime("%m-%d-%Y")
	output_file = "ORDERS {}.xlsx".format(today)
	star_file = "STAR {}.xlsx".format(today)
	sbw_file = "SBW {}.xlsx".format(today)
	dir_path = os.path.join(os.getcwd(), 'Output Sheets')
	os.makedirs(dir_path, exist_ok=True)
	output_wb.save(os.path.join(dir_path, output_file))

	# Splitting into separate Star Interactive and SBW Sheets
	print('Splitting into Star Interactive and SBW Sheets...')
	
	for row in range(2, output_sheet.max_row + 1):
		
		order_sku = output_sheet['CR' + str(row)].value
		star_check = cur.execute("SELECT 1 FROM star_inventory WHERE item_sku = ?", (order_sku,)).fetchall()
		sbw_check = cur.execute("SELECT 1 FROM sbw_inventory WHERE item_sku = ?", (order_sku,)).fetchall()

		if star_check and sbw_check:
			print("{} exists in both STAR and SBW tables - sorting based off of UPC code.".format(order_sku))
			order_upc = output_sheet['CQ' + str(row)].value
			star_upc_check = cur.execute("SELECT 1 FROM star_inventory WHERE item_upc = ?", (order_upc,)).fetchall()
			if star_upc_check:
				sbw_check = False
				print('UPC found in STAR table - going with STAR SKU')
			else:
				star_check = False
				print('UPC not found in STAR table - defaulting to SBW SKU')

		if star_check:	
			for col in range(1, output_sheet.max_column + 1):
				col_letter = openpyxl.cell.cell.get_column_letter(col)
				star_sheet[col_letter + str(star_row)] = output_sheet[col_letter + str(row)].value
			star_row += 1

		elif sbw_check:
			for col in range(1, output_sheet.max_column + 1):
				col_letter = openpyxl.cell.cell.get_column_letter(col)
				sbw_sheet[col_letter + str(sbw_row)] = output_sheet[col_letter + str(row)].value
			sbw_row += 1
			
		else:
			print("WARNING: {} NOT FOUND IN DATABASE".format(order_sku))

	if star_sheet.max_row > 1:
		star_wb.save(os.path.join(dir_path, star_file))
		print('Star Interactive sheet writen.')
	
	if sbw_sheet.max_row > 1:
		sbw_wb.save(os.path.join(dir_path, sbw_file))
		print('SBW sheet writen.')
	
	print('Done.')
	print('Total SKUs:', output_sheet.max_row - 1)
	print('Star Interactive SKUs', star_sheet.max_row - 1)
	print('SBW SKUs', sbw_sheet.max_row - 1)
	print('-----')
	for row in error_rows:
		print('WARNING! Potential error on row', row)

	cur.close()
	conn.close()