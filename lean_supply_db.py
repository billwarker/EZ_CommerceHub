import openpyxl
import pymysql

print('Starting...')

conn = pymysql.connect(host='127.0.0.1', user='root', passwd='Greengiant90',
						db='mysql', charset='utf8')
cur = conn.cursor()
cur.execute("USE star_interactive")

inventory_file = "lean supply inventory.xlsx" 
wb = openpyxl.load_workbook(inventory_file)
sheet = wb.active

total_items = sheet.max_row - 1
added_items = 0

added_skus = set()
failed_skus = list()
no_upc = set()

for row in range(2, sheet.max_row + 1):
	try:
		sku = sheet['B' + str(row)].value
		item_desc = sheet['C' + str(row)].value
		upc = sheet['I' + str(row)].value
		if upc == '':
			upc = str(0)
			no_upc.add(sku)
		if sku not in added_skus:
			query = """INSERT INTO lean_supply (sku, item_desc, upc)
					 VALUES (%s, %s, %s)"""

			cur.execute(query, (sku, item_desc, upc))
			conn.commit()
			print('Added', sku)
			added_skus.add(sku)
			added_items += 1
	except Exception as e:
		print(e)
		failed = (sku, e)
		failed_skus.append(failed)


# Close MySQL Database
cur.close()
conn.close()
print('Done, {} out of {} total items successfully added to database.'.format(added_items, total_items))
print('Number of SKUs with no associated UPC: ', str(len(no_upc)))
print('Failed SKUs:')
if failed_skus > 0:
	for sku in failed_skus:
		print(sku[0], sku[1])





